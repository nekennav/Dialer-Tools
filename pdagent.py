import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import base64

# Set page configuration
st.set_page_config(page_title="PREDICTIVE SUMMARIZER", page_icon="📊", layout="wide")

# Custom CSS for Demon Slayer-themed styling
background_image = "url('https://images7.alphacoders.com/139/1398431.jpg')"
st.markdown(
    f"""
    <style>
    /* Main app background with Demon Slayer image */
    .stApp {{
        background-image: {background_image};
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
        background-repeat: no-repeat;
        color: #f0f0f0; /* Light gray for text readability */
        background-color: #2c1810; /* Dark brown fallback color */
    }}
    /* Semi-transparent overlay for content readability */
    .main-container {{
        background: rgba(44, 24, 16, 0.8); /* Dark brown semi-transparent overlay */
        padding: 30px;
        border-radius: 15px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.5);
        margin: 20px;
    }}
    /* Title styling with Demon Slayer-inspired colors */
    h1 {{
        color: #ffffff; /* White for PREDICTIVE SUMMARIZER */
        font-family: 'Arial', sans-serif;
        text-align: center;
        text-shadow: 2px 2px 6px rgba(0, 0, 0, 0.7);
    }}
    /* Subtitle styling */
    h3 {{
        color: #1e90ff; /* Blue for NEKENNAV */
        font-family: 'Arial', sans-serif;
        text-align: center;
    }}
    /* File uploader styling */
    .stFileUploader {{
        background-color: rgba(255, 107, 53, 0.1); /* Light orange tint */
        border: 2px solid #ff6b35; /* Solid orange border */
        border-radius: 10px;
        padding: 10px;
    }}
    /* Button styling */
    .stButton > button {{
        background-color: #ff6b35; /* Vibrant orange */
        color: #f0f0f0; /* Light gray text */
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: bold;
        transition: background-color 0.3s;
        border: 1px solid #ffd23f; /* Golden border */
    }}
    .stButton > button:hover {{
        background-color: #e55a2b; /* Darker orange on hover */
        border-color: #ff6b35; /* Orange on hover */
    }}
    /* Text area and dataframe styling */
    .stTextArea textarea, .stDataFrame {{
        background-color: rgba(255, 107, 53, 0.15); /* Subtle orange tint */
        color: #f0f0f0; /* Light gray text */
        border: 1px solid #ff6b35; /* Solid orange border */
        border-radius: 8px;
    }}
    /* Alert styling */
    .stAlert {{
        background-color: rgba(255, 211, 63, 0.2); /* Light golden tint */
        color: #f0f0f0; /* Light gray text */
        border-radius: 8px;
    }}
    /* Table text */
    .stText, table {{
        color: #f0f0f0; /* Light gray for table text */
        font-family: 'Arial', sans-serif;
    }}
    </style>
    """,
    unsafe_allow_html=True
)

# Title and description
st.markdown('<div class="main-container">', unsafe_allow_html=True)
st.title("PREDICTIVE SUMMARIZER")
st.markdown("<h3>NEKENNAV</h3>", unsafe_allow_html=True)

# File uploader widget
uploaded_files = st.file_uploader(
    "Choose Excel files to merge",
    accept_multiple_files=True,
    type=['xlsx', 'xls']
)

# Create a directory to store uploaded files temporarily
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# Function to convert time to seconds
def time_to_seconds(time_val):
    try:
        if pd.isna(time_val):
            return 0
        if isinstance(time_val, (int, float)):
            return float(time_val)
        if isinstance(time_val, str):
            parts = time_val.split(':')
            parts = [p.strip() for p in parts]
            if len(parts) == 3: # HH:MM:SS
                h, m, s = map(int, parts)
                return h * 3600 + m * 60 + s
            elif len(parts) == 2: # MM:SS
                m, s = map(int, parts)
                return m * 60 + s
        return 0
    except:
        return 0

# Function to format seconds to [h]:mm:ss for display
def seconds_to_time(seconds):
    if pd.isna(seconds):
        return "0:00:00"
    seconds = int(seconds)
    hours = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
    return f"{hours}:{minutes:02d}:{seconds:02d}"

# Function to merge and aggregate Excel files
def merge_excel_files(files):
    try:
        dfs = []
        for file in files:
            file_path = os.path.join(UPLOAD_DIR, file.name)
            with open(file_path, "wb") as f:
                f.write(file.getbuffer())
            df = pd.read_excel(file_path, engine=None)
            dfs.append(df)
        
        if dfs:
            merged_df = pd.concat(dfs, ignore_index=True)
            columns_to_drop = ['SNo.', 'Total Calls', 'Pause Count']
            merged_df = merged_df.drop(columns=[col for col in columns_to_drop if col in merged_df.columns])
            
            if 'Collector Name' in merged_df.columns:
                merged_df = merged_df[merged_df['Collector Name'].notna() & (merged_df['Collector Name'].str.strip() != '')]
            else:
                return None, "Collector Name column not found in the data."
            
            time_columns = [
                'Spent Time', 'Talk Time', 'AVG Talk Time', 'Wait Time',
                'Average Wait Time', 'Write Time', 'AVG Write Time', 'Pause Time'
            ]
            valid_time_columns = [col for col in time_columns if col in merged_df.columns]
            
            for col in valid_time_columns:
                merged_df[col] = merged_df[col].apply(time_to_seconds)
            
            agg_dict = {col: 'sum' for col in valid_time_columns}
            other_columns = [col for col in merged_df.columns if col not in valid_time_columns + ['Collector Name']]
            for col in other_columns:
                agg_dict[col] = 'first'
            
            if not agg_dict:
                return merged_df, None
            
            merged_df = merged_df.groupby('Collector Name').agg(agg_dict).reset_index()
            
            avg_row = {'Collector Name': 'Average'}
            for col in valid_time_columns:
                avg_row[col] = merged_df[col].mean()
            for col in other_columns:
                avg_row[col] = None
            
            avg_df = pd.DataFrame([avg_row])
            merged_df = pd.concat([merged_df, avg_df], ignore_index=True)
            
            return merged_df, None
        else:
            return None, "No valid Excel files uploaded."
    
    except Exception as e:
        return None, f"Error merging files: {str(e)}"

# Process uploaded files
if uploaded_files:
    with st.spinner("Merging Excel files..."):
        merged_df, error = merge_excel_files(uploaded_files)
    
    if error:
        st.error(error)
    else:
        st.success(f"Successfully uploaded {len(uploaded_files)} file(s)!")
        st.write("**Preview of Merged Data**")
        display_df = merged_df.copy()
        time_columns = [
            'Spent Time', 'Talk Time', 'AVG Talk Time', 'Wait Time',
            'Average Wait Time', 'Write Time', 'AVG Write Time', 'Pause Time'
        ]
        valid_time_columns = [col for col in time_columns if col in display_df.columns]
        for col in valid_time_columns:
            display_df[col] = display_df[col].apply(seconds_to_time)
        
        st.dataframe(display_df, use_container_width=True)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            excel_df = merged_df.copy()
            for col in valid_time_columns:
                excel_df[col] = excel_df[col] / 86400.0
            
            excel_df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            for col in valid_time_columns:
                col_idx = merged_df.columns.get_loc(col) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(merged_df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = '[h]:mm:ss'
                    cell.alignment = Alignment(horizontal='right')
                
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.alignment = Alignment(horizontal='right')
            
            for col_idx in range(1, len(merged_df.columns) + 1):
                col_name = merged_df.columns[col_idx - 1]
                if col_name != 'Collector Name':
                    col_letter = get_column_letter(col_idx)
                    for row in range(1, len(merged_df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        if col_name not in valid_time_columns:
                            cell.alignment = Alignment(horizontal='right')
        
        output.seek(0)
        
        st.download_button(
            label="Download Merged Excel File",
            data=output,
            file_name=f"Merged_Excel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_merged"
        )
else:
    st.info("Please upload one or more Excel files to merge.")

# Clean up temporary upload directory
import shutil
if os.path.exists(UPLOAD_DIR):
    shutil.rmtree(UPLOAD_DIR)

st.markdown('</div>', unsafe_allow_html=True)
